from math import sqrt
from os import path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference

from client.mlcar import logger

log = logger.Logger("data")


class DataLogger(object):
    def __init__(self, file_base):
        self._fb = file_base
        self._wb = None
        self._s = None
        self._c_row = 0
        self._gen = 0
        self._trial = 0

    def set_gen(self, gen):
        file = self._fb % gen
        if path.exists(file):
            log.info("The workbook exists! Loading...")
            self._wb = load_workbook(filename=file)
        else:
            log.info("Creating the new workbook")
            self._wb = Workbook()
        self._s = self._wb.active
        self._s.title = "gen_%d" % gen
        self._c_row = 0
        self._gen = gen

    def set_trial(self, t_num):
        if ("t_%d" % t_num) in self._wb.get_sheet_names():
            t_s = self._wb.get_sheet_by_name("t_%d" % t_num)
            self._wb.remove_sheet(t_s)
        self._trial = t_num
        self._s = self._wb.create_sheet(title="t_%d" % t_num)
        self._s.append(["Time stamp", "Viewable", "Current Error", "Current Absolute Error", "Future Error",
                        "Future Absolute Error", "Power", "Absolute Power", "Turn", "Absolute Turn", "Total Time",
                        "Total Frames", "Average Error", "Lowest Error", "Highest Error", "Average Power",
                        "Lowest Power", "Highest Power", "Average Turn", "Lowest Turn", "Highest Turn",
                        "Completed Course"])

    def get_average(self, data, absolute=False):
        if absolute:
            data = map(abs, data)
        try:
            return float(sum(data)) / float(len(data))
        except:
            return "No data available"

    def get_median(self, data, absolute=False):
        if absolute:
            data = map(abs, data)
        try:
            n = len(data)
            if n == 0:
                raise Exception("No data")
            if n % 2 == 1:
                return sorted(data)[n // 2]
            else:
                return sum(sorted(data)[n // 2 - 1:n // 2 + 1]) / 2.0
        except:
            return "No data available"

    def get_peaks(self, data, absolute=False):
        if absolute:
            data = map(abs, data)
        if len(data) == 0:
            return "No data available"
        lowest = 100000000
        highest = -10000000
        for d in data:
            if d < lowest:
                lowest = d
            if d > highest:
                highest = d
        return lowest, highest

    def get_standard_deviation(self, data):
        try:
            n = len(data)
            if n == 0:
                raise Exception("No data")
            mean = self.get_average(data)
            return sqrt(sum([((d - mean) ** 2) for d in data]) / (n - 1))
        except:
            return "No data available"

    def add_data(self, data):
        self._s.append(data)
        self._c_row += 1

    def get_data(self, row):
        return list([d for d in [d.value for d in list(self._s.rows)[row + 1]] if d is not None])

    def get_v_data(self, column):
        data = []
        for d in self._s[column]:
            try:
                data.append(float(d.value))
            except:
                pass
        return data

    def set_data(self, row, data):
        for col, val in enumerate(data, start=1):
            self._s.cell(row=row, column=col).value = val

    def get_data_count(self):
        return self._c_row

    def create_line_graph(self, title, location, y_axis="Proportions", col_ref=("B",)):
        l = LineChart()
        l.title = title
        l.style = 13
        l.x_axis.title = "Time Stamp"
        l.y_axis.title = y_axis

        for c in col_ref:
            data = Reference(self._s, range_string="t_%d!%s1:%s%d" % (self._trial, c, c, self._c_row + 1))
            l.add_data(data, titles_from_data=True)

        colors = ["009688", "03A9F4", "673AB7", "FFC107", "FF5722", "607D8B"]

        for i in range(0, len(list(l.series))):
            pass

        time_stamps = Reference(self._s, range_string="t_%d!A2:A%d" % (self._trial, self._c_row + 1))
        l.set_categories(time_stamps)
        l.width = 30
        l.height = 15

        self._s.add_chart(l, location)

    def save_data(self):
        if self._wb is None or self._s is None:
            return

        # Fixes all of the column widths
        dims = {}
        for row in self._s.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len("%s" % cell.value)))
        for col, value in dims.items():
            self._s.column_dimensions[col].width = value

        # Save the file
        self._wb.save(self._fb % self._gen)
