from os import path

from openpyxl import Workbook, load_workbook

from .. import logger

log = logger.Logger("data")


class DataLogger(object):
    def __init__(self, file_base):
        self._fb = file_base
        self._wb = None
        self._s = None
        self._c_row = 0
        self._gen = 0

    def set_gen(self, gen):
        file = self._fb % gen
        if path.exists(file):
            log.info("The workbook exists! Loading...")
            self._wb = load_workbook(filename=file)
        else:
            log.info("Creating the new workbook")
            self._wb = Workbook()
        self._s = self._wb.active
        self._c_row = 0
        self._gen = gen

    def set_trial(self, t_num):
        if ("t_%d" % t_num) in self._wb.get_sheet_names():
            self._s = self._wb.get_sheet_by_name("t_%d" % t_num)
            try:
                for row in self._wb.rows:
                    for cell in row:
                        cell.value = None
            except:
                log.error("FAILED TO CLEAR THE SHEET MALFORMED DATA!")
        else:
            self._s = self._wb.create_sheet(title="t_%d" % t_num)
        self._s.append(["Timestamp", "Viewable", "Current Error", "Future Error", "Power", "Turn", "Total Time",
                        "Total Frames", "Average Error", "Lowest Error", "Highest Error", "Average Power",
                        "Lowest Power", "Highest Power", "Completed Course"])

    def add_data(self, data):
        self._s.append(data)

    def save_data(self):
        if self._wb is None or self._s is None:
            return

        # Fixes all of the column widths
        dims = {}
        for row in self._wb.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        for col, value in dims.items():
            self._wb.column_dimensions[col].width = value

        # Save the file
        self._wb.save(self._fb % self._gen)
